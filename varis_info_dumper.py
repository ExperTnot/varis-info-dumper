import os
import re
import time
import sys
import re
from docx import Document
import openpyxl
from tkinter import Tk, Button, Label, Frame
import threading
from colorama import init, Fore, Style

CONFIG_FILE_DOCX = "config.txt"
CONFIG_FILE_XLSX = "config_xlsx.txt"

YES_CHOICES = ['yes', 'y']
NO_CHOICES = ['no', 'n']

COLUMN_INDEX_ID = 3 # Column D is index 3
COLUMN_INDEX_GENDER = 10 # Column J is index 10

#Threading
def gui_thread(lines):
    r = Tk()
    r.title("Copy Lines to Clipboard")

    def copy_to_clipboard(line_to_copy):
        r.clipboard_clear()
        r.clipboard_append(line_to_copy)
        r.update()

    def close_window():
        r.quit()
        r.destroy()
    
    background_colors = ["white", "lightgrey", "darkgrey"]  # Define background colors
    row_num = 0  # Initialize row number

    for line_num, line in enumerate(lines, start=1):
        line = line.strip()
        frame = Frame(r, bg=background_colors[row_num % len(background_colors)])
        label_text = f"{line_num}. {line}"
        label = Label(frame, text=label_text, bg=background_colors[row_num % len(background_colors)])
        copy_button = Button(frame, text=f"Copy {line_num}", command=lambda line=line: copy_to_clipboard(line))
        label.pack(side="left")
        copy_button.pack(side="right")
        frame.pack(fill="both", expand=True)
        row_num += 1

    close_button = Button(r, text="Close", command=close_window)
    close_button.pack()

    r.protocol("WM_DELETE_WINDOW", close_window)
    
    window_height = len(lines) * 50  # Adjust the multiplier as needed
    r.geometry(f"200x{window_height}") # Adjust the width as needed
    r.mainloop()

# Function to extract data from a Word document
def extract_word_data(docx_file):
    document = Document(docx_file)
    data = []

    for paragraph in document.paragraphs:
        text = paragraph.text

        # Use regular expression to find HP numbers
        hp_numbers = re.findall(r'HP:\d{7}', text)

        if hp_numbers:
            data.extend(hp_numbers)

    return data

# Function to open a file with a 1-second delay
def open_file_with_delay(file_path):
    time.sleep(1)
    os.system(f"start {file_path}")

# Function to get the folder path from the user and save it to the configuration file
def get_folder_path():
    folder_path = input("Enter the folder path where your .docx files are located: ")
    with open(CONFIG_FILE_DOCX, "w") as config_file:
        config_file.write(folder_path)
    return folder_path

# Function to read the folder path from the configuration file
def read_folder_path():
    if os.path.exists(CONFIG_FILE_DOCX):
        with open(CONFIG_FILE_DOCX, "r") as config_file:
            return config_file.read()
    else:
        return None

def get_cell_value(row, column_index):
    """Helper function to get cell value from a row"""
    if len(row) > column_index:
        return row[column_index]
    print(f"Column index {column_index} is out of range.")
    return None

def search_next_rows(sheet, start_row, column_index, search_values):
    """Helper function to search for specific values in next rows"""
    for row_number in range(start_row, sheet.max_row + 1):
        cell_value = sheet.cell(row=row_number, column=column_index).value
        if cell_value in search_values:
            return (cell_value, get_cell_value(sheet[row_number], COLUMN_INDEX_ID).value)
    print(f"Values {search_values} not found in Column {column_index}.")
    return (None, None)

def search_excel_and_extract_data(excel_file, search_value):
    wb = openpyxl.load_workbook(excel_file)
    sheet = wb.active

    extracted_data = None

    # Iterate through rows in Column D (ID) to find the search value
    for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        cell_value = str(get_cell_value(row, COLUMN_INDEX_ID))

        if search_value in cell_value:
            # Extract data from Column J (if available)
            column_j_value = get_cell_value(row, COLUMN_INDEX_GENDER - 1) # -1 because here array -> 0 based index

            extracted_data = (cell_value, column_j_value)

            # Debug print after finding a matching cell in Column C
            print(f"Found {Fore.GREEN}{search_value}{Style.RESET_ALL} in cell {Fore.GREEN}{cell_value}{Style.RESET_ALL}, extracted Column I value: {Fore.GREEN}{column_j_value}{Style.RESET_ALL}")

            # Search for "Vater" or "Mutter" in the next rows
            next_row_values = search_next_rows(sheet, row_number + 1, COLUMN_INDEX_GENDER, ["Vater", "Mutter"]) 
                
            if next_row_values[0]:   
                extracted_data += next_row_values

                # Debug print after finding "Vater" or "Mutter" in the next rows
                print(f"Found {Fore.GREEN}{next_row_values[0]}{Style.RESET_ALL} in the next row, Column C: {Fore.GREEN}{next_row_values[1]}{Style.RESET_ALL}")

                # If the first check was successful, check a second time before stopping
                next_row_values_second_check = search_next_rows(sheet, row_number + 2, COLUMN_INDEX_GENDER, ["Vater", "Mutter"])
                
                if next_row_values_second_check[0]:
                    # If the second check is successful, update the extracted data and stop searching
                    extracted_data += next_row_values_second_check

                    print(f"Found {Fore.GREEN}{next_row_values_second_check[0]}{Style.RESET_ALL} in the subsequent row, Column C: {Fore.GREEN}{next_row_values_second_check[1]}{Style.RESET_ALL}")
                
            break  # Stop searching after finding the first match

    wb.close()
    return extracted_data

# Function to get the Excel file path from the user and save it to the configuration file
def get_excel_file_path():
    excel_file_path = input("Enter the path to the Excel sheet (xlsx file): ")
    with open(CONFIG_FILE_XLSX, "w") as config_file:
        config_file.write(excel_file_path)
    return excel_file_path

# Function to read the Excel file path from the configuration file
def read_excel_file_path():
    if os.path.exists(CONFIG_FILE_XLSX):
        with open(CONFIG_FILE_XLSX, "r") as config_file:
            return config_file.read()
    else:
        return None

# Function to add data to a text file
def add_data_to_text_file(file_path, extracted_data):
    with open(file_path, "a") as file:
        if extracted_data is not None:
            for item in extracted_data[1:]:  # Start from the second item
                if item is not None:
                    file.write(str(item) + "\n")  # Append each item with a newline character

def read_config_file(file_type):
    file_path = read_folder_path() if file_type == 'word' else read_excel_file_path()
    if not file_path:
        print(f"Configuration file for {file_type} documents not found or {file_type} file path not configured.")
        file_path = get_folder_path() if file_type == 'word' else get_excel_file_path()
    return file_path

def main():
    init() # Initialize colorama
    while True:
        # Check if the configuration files exist
        folder_path = read_config_file('word')
        excel_file_path = read_config_file('excel')

        if not os.path.exists(excel_file_path):
            print(f"Excel file '{excel_file_path}' not found.")
            return

        # Input: Provide a partial number to search for
        partial_number = input("Enter a partial number to search for: ")

        folder_dir = folder_path # Path to the folder with the Word documents

        # Search for folders with the same leading number as the partial number
        matching_folders = []
        for folder_name in os.listdir(folder_dir):
            if os.path.isdir(os.path.join(folder_dir, folder_name)) and folder_name.startswith(partial_number):
                matching_folders.append(folder_name)

        # Determine the directory of the executable (script or .exe)
        if getattr(sys, 'frozen', False):
        # The script is running as a compiled executable (.exe)
            script_dir = os.path.dirname(sys.executable)
        else:
        # The script is running as a regular Python script
            script_dir = os.path.dirname(__file__)

        # Check if any matching folders were found
        if matching_folders:

            # Iterate through the matching folders and extract data from the .docx files
            for folder_name in matching_folders:
                folder_path = os.path.join(folder_dir, folder_name)
                docx_files = [file for file in os.listdir(folder_path) if file.endswith(".docx") and file.startswith(folder_name)]

                if not docx_files:
                    print(f"No .docx files with the same leading number found in folder '{folder_name}'.")
                    continue

                docx_file = os.path.join(folder_path, docx_files[0])  # Use the first .docx file found

                word_data = extract_word_data(docx_file) # Extract data from the .docx file

                # Path for the output text file in the same directory as the script
                output_file_path = os.path.join(script_dir, f"{folder_name}.txt")

                # Write the collected HP numbers to the output text file
                with open(output_file_path, "w") as output_file:
                    for hp_number in word_data:
                        output_file.write(hp_number + "\n")

                print(f"Data from {Fore.GREEN}{docx_files[0]}{Style.RESET_ALL} has been saved to {Fore.GREEN}{folder_name}.txt{Style.RESET_ALL}")
                
            # Input: Provide a 4-digit number to search for in the Excel sheet (xlsx file)
            search_value = partial_number #input("Enter a 4-digit number to search for in the Excel sheet: ")

            # Search the Excel sheet and extract data
            extracted_data = search_excel_and_extract_data(excel_file_path, search_value)

            if extracted_data is None:
                print(f"No data found for '{search_value}' in the Excel sheet.")
                return

            # Display the extracted data
            cell_value = extracted_data[0] if extracted_data is not None else None

            if cell_value is not None:
                print(f"Cell Value: {Fore.GREEN}{cell_value}{Style.RESET_ALL}")
            else:
                print(f"No data found for '{search_value}' in the Excel sheet.")

            # Create a text file with the cell value as the filename and add the data to the file
            text_file_path = os.path.join(script_dir, f"{cell_value}.txt")
            add_data_to_text_file(text_file_path, extracted_data)
                        
            # Start thread to read file and display gui
            if os.path.exists(output_file_path):
                with open(output_file_path, "r") as output_file:
                    lines = output_file.readlines()
                    if lines:
                        # Determine the required window height based on the number of lines
                        window_height = len(lines) * 50

                        # Create a separate thread to run the GUI
                        gui_thread_thread = threading.Thread(target=gui_thread, args=(lines,), daemon=True)
                        gui_thread_thread.start()
            
        else:
            print(f"No matching folders found for '{partial_number}' in Word documents.")
    
        exit_program = False # Flag variable
        print(f"{Fore.RED}DONT FORGET TO CLOSE THE GUI IF IT IS STILL OPEN!{Style.RESET_ALL}")
        while True: # Ask for next or close
            confirmation = input(f"Do you want to exit the program? {Fore.BLUE}(y/n){Style.RESET_ALL}: ")
            if confirmation.lower() in YES_CHOICES:
                print("Exiting the program.")
                exit_program = True # Set the flag to True
                break
            elif confirmation.lower() in NO_CHOICES:
                print("------------------------------------------------------------------------")
                break
            else:
                print(f"{Fore.RED}Invalid input.{Style.RESET_ALL}")
        
        if exit_program: # Check the flag
            break
            

if __name__ == "__main__":
    main()
